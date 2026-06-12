import openpyxl
from datetime import datetime, date

src = r"C:\Tiim Market Base Loyihasi\sotuv_excel.xlsx"
dst = r"C:\Tiim Market Base Loyihasi\sotuv_hafta_test.xlsx"

wb_in = openpyxl.load_workbook(src, read_only=True, data_only=True)
sh_in = wb_in.active
sh_in.reset_dimensions()

wb_out = openpyxl.Workbook()
sh_out = wb_out.active

rows = sh_in.iter_rows(values_only=True)
headers = list(next(rows))
sh_out.append(headers)

date_col = headers.index("Date")
count = 0
for row in rows:
    dv = row[date_col]
    if isinstance(dv, datetime): sd = dv.date()
    elif isinstance(dv, date): sd = dv
    else:
        try: sd = datetime.fromisoformat(str(dv)[:10]).date()
        except: continue
    if sd.month == 5 and 1 <= sd.day <= 7:
        sh_out.append(list(row))
        count += 1

wb_in.close()
wb_out.save(dst)
print(f"Tayyor: {count} satr (May 1-7) -> sotuv_hafta_test.xlsx")
print(f"Asl fayl: TEGILMADI")
