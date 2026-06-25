#!/usr/bin/env python3
"""
import_arrivals.py — Invan'dan qo'lda export qilingan "Поставки (по товарам)"
(Inventory purchase, AllOrderItemsReport) Excel faylidan har bir SKU uchun
oxirgi KIRIM (ReceivedDate, Status=Received) sanasini chiqarib, arrival_data.json
ga yozadi.

Bu fayl avtomatik 30-daqiqalik sync'ga kirmaydi (Invan API'da bunday endpoint yo'q),
shu sabab vaqti-vaqti bilan qo'lda: Invan > Reports > Inventory > Inventory purchase
> Export qilib, shu skriptni qayta ishga tushirish kerak.

Foydalanish:
    python import_arrivals.py "C:/Users/User/Downloads/Поставки (по товарам).xlsx"
"""
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent


def main():
    if len(sys.argv) < 2:
        sys.exit("Foydalanish: python import_arrivals.py <excel_fayl_yoli>")
    src = Path(sys.argv[1])
    if not src.exists():
        sys.exit(f"Fayl topilmadi: {src}")

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["AllOrderItemsReport"] if "AllOrderItemsReport" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h or "").strip() for h in rows[0]]
    idx = {h: i for i, h in enumerate(header)}

    sku_i = idx.get("SKU")
    received_i = idx.get("ReceivedDate")
    status_i = idx.get("Status")
    qty_i = idx.get("Quantity")
    if sku_i is None or received_i is None or status_i is None:
        sys.exit(f"Kutilgan ustunlar topilmadi. Mavjud ustunlar: {header}")

    last_arrival = {}
    matched = 0
    for row in rows[1:]:
        if not row or row[status_i] != "Received":
            continue
        sku = str(row[sku_i] or "").strip()
        received = row[received_i]
        if not sku or not received:
            continue
        received_s = str(received)[:10]
        qty = row[qty_i] if qty_i is not None else None
        if sku not in last_arrival or received_s > last_arrival[sku]["date"]:
            last_arrival[sku] = {"date": received_s, "qty": qty}
        matched += 1

    out_path = ROOT / "arrival_data.json"
    out_path.write_text(json.dumps(last_arrival, ensure_ascii=False), encoding="utf-8")
    print(f"{len(last_arrival):,} ta SKU uchun oxirgi kirim sanasi topildi ({matched:,} 'Received' qator tahlil qilindi).")
    print(f"Saqlandi: {out_path}")


if __name__ == "__main__":
    main()
