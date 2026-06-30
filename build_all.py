#!/usr/bin/env python3
"""
build_all.py — Yangi oy ma'lumotlari bilan sales.html ni qayta qurish.

Foydalanish:
    python build_all.py --sales sotuv.xlsx --products tovarlar.xlsx
"""
import json, math, re, argparse, sys, shutil
from collections import defaultdict, Counter
from datetime import datetime, date
from pathlib import Path

ROOT = Path(__file__).parent

# ─── ulgurji aniqlash konstantalari ───
DUMMY_TINS = {"", "999999999", "888888888", "555555555"}
BUSINESS_WORDS = (
    "mchj", "ooo", "ооо", "xk", "hotel", "kafe", "cafe", "oshxona",
    "restaurant", "restoran", "supply", "development", "market", "магазин",
    "school", "maktab", "bank", "aj ",
)
NON_BUSINESS = {"", "xodimlar", "nujda", "farrux"}

WEEKDAYS_UZ    = ["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba", "Yakshanba"]
MONTHS_UZ      = {1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel", 5: "May", 6: "Iyun",
                  7: "Iyul", 8: "Avgust", 9: "Sentabr", 10: "Oktabr", 11: "Noyabr", 12: "Dekabr"}


# ─── yordamchi funksiyalar ───
def norm(v):
    return re.sub(r"\s+", " ", str(v or "")).strip().lower()

def rq(v):
    v = float(v)
    return int(round(v)) if abs(v - round(v)) < 0.001 else round(v, 3)

def pctl(vals, p):
    if not vals: return 0.0
    s = sorted(vals)
    return float(s[min(len(s) - 1, int(p * (len(s) - 1)))])

def median(vals): return pctl(vals, 0.5)

def is_wholesale(customer, tin):
    cn = norm(customer)
    tn = str(tin or "").strip()
    if tn not in DUMMY_TINS and re.fullmatch(r"\d{9}", tn):
        return True
    if cn in NON_BUSINESS:
        return False
    return any(w in cn for w in BUSINESS_WORDS)


# ─── tovarlar excel o'qish ───
def read_products(path):
    """Tovarlar Excel → {sku: {name, cat, sub, tp, su, p, a}}"""
    try:
        import openpyxl
    except ImportError:
        print("Xato: openpyxl o'rnatilmagan. pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = wb.active
    sh.reset_dimensions()
    rows = sh.iter_rows(values_only=True)
    headers = [str(h or "").strip() for h in next(rows)]
    idx = {h.lower(): i for i, h in enumerate(headers)}

    def find(*names):
        for n in names:
            nl = n.lower()
            for h, i in idx.items():
                if nl in h:
                    return i
        return -1

    c_sku   = find("sku")
    c_name  = find("name")
    c_cat   = find("category")
    c_sub   = find("sub_category", "subcategory")
    c_type  = find("type")
    c_sup   = find("supplier")
    c_price = find("retail_price")
    c_amt   = find("amount")

    products = {}
    for row in rows:
        sku = str(row[c_sku] or "").strip() if c_sku >= 0 else ""
        if not sku:
            continue
        products[sku] = {
            "name": norm(row[c_name]) if c_name >= 0 else "",
            "cat":  str(row[c_cat]  or "").strip() if c_cat  >= 0 else "",
            "sub":  str(row[c_sub]  or "").strip() if c_sub  >= 0 else "",
            "tp":   str(row[c_type] or "").strip() if c_type >= 0 else "",
            "su":   str(row[c_sup]  or "").strip() if c_sup  >= 0 else "",
            "p":    float(row[c_price] or 0)       if c_price >= 0 else 0.0,
            "a":    float(row[c_amt]   or 0)       if c_amt  >= 0 else 0.0,
        }
    wb.close()
    return products


# ─── sotuv excel o'qish ───
def read_sales(path):
    """Sotuv Excel → receipts, pnames, pskus, min_date, max_date"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = wb.active
    sh.reset_dimensions()
    rows = sh.iter_rows(values_only=True)
    headers = [str(h or "").strip() for h in next(rows)]
    idx = {h: i for i, h in enumerate(headers)}

    missing = [c for c in ("Date", "Receipt#", "Sku", "Item", "Qty", "Total Price") if c not in idx]
    if missing:
        raise ValueError(f"Sotuv Excel da ustunlar topilmadi: {', '.join(missing)}")

    receipts = {}
    pnames = defaultdict(Counter)
    pskus  = {}
    pcats  = defaultdict(Counter)   # pk → kategoriya (sotuvdagi Category ustuni)
    refund_total  = 0.0             # qaytarilgan summa
    refund_by_day = defaultdict(float)  # sana → qaytarish summasi
    min_d = max_d = None
    proc = 0

    for row in rows:
        sale_type = norm(row[idx["Sale Type"]]) if "Sale Type" in idx else "sale"
        qty = float(row[idx["Qty"]] or 0)
        total_p = float(row[idx["Total Price"]] or 0)

        # sanani oldindan ajratib olamiz (qaytarish uchun ham kerak)
        dv = row[idx["Date"]]
        if isinstance(dv, datetime): sd = dv.date()
        elif isinstance(dv, date):   sd = dv
        else:
            try:    sd = datetime.fromisoformat(str(dv)[:10]).date()
            except: sd = None

        # qaytarishlarni alohida yig'amiz (refund yoki manfiy miqdor)
        if sale_type != "sale" or qty < 0:
            refund_total += abs(total_p)
            if sd is not None:
                refund_by_day[sd] += abs(total_p)
            continue
        if qty <= 0 or sd is None:
            continue

        min_d = sd if min_d is None or sd < min_d else min_d
        max_d = sd if max_d is None or sd > max_d else max_d

        rid = str(row[idx["Receipt#"]] or "").strip()
        if not rid:
            continue

        customer = str(row[idx["Customer"]] or "").strip() if "Customer" in idx else ""
        tin      = str(row[idx["TIN"]]      or "").strip() if "TIN"      in idx else ""
        employee = str(row[idx["Employee"]] or "").strip() if "Employee" in idx else ""
        category = str(row[idx["Category"]] or "").strip() if "Category" in idx else ""

        rc = receipts.setdefault(rid, {
            "date": sd, "customer": "", "tin": "", "employee": "",
            "items": defaultdict(float), "item_rev": defaultdict(float),
        })
        if customer: rc["customer"] = customer
        if tin:      rc["tin"]      = tin
        if employee: rc["employee"] = employee

        sku  = str(row[idx["Sku"]] or "").strip()
        name = norm(row[idx["Item"]])
        pk   = "sku:" + sku if sku else "name:" + name
        pnames[pk][name] += 1
        pskus[pk] = sku
        if category:
            pcats[pk][category] += 1
        rc["items"][pk]    += qty
        rc["item_rev"][pk] += total_p

        proc += 1
        if proc % 100_000 == 0:
            print(f"      ...{proc:,} satr o'qildi")

    wb.close()
    return receipts, pnames, pskus, pcats, refund_total, refund_by_day, min_d, max_d


# ─── ulgurji qoidalar ───
def compute_rules(receipts):
    anon = defaultdict(list)
    for rc in receipts.values():
        if is_wholesale(rc["customer"], rc["tin"]):
            continue
        for pk, qty in rc["items"].items():
            anon[pk].append(qty)

    rules = {}
    for pk, qtys in anon.items():
        m    = median(qtys)
        sig  = median([abs(q - m) for q in qtys]) * 1.4826
        uf   = 3 if any(abs(q - round(q)) > 0.001 for q in qtys) else 5
        pre  = max(uf, m * 3, m + 6 * sig)
        rs   = [q for q in qtys if q <= pre] or qtys
        rm   = median(rs)
        rsig = median([abs(q - rm) for q in rs]) * 1.4826
        rp90 = pctl(rs, 0.90)
        rp95 = pctl(rs, 0.95)
        cap  = max(uf, rp95, rm + 4 * rsig)
        thr  = max(cap + uf, cap * 2, rp90 * 3, rm + 8 * rsig)
        rules[pk] = {"cap": cap, "thr": thr, "med": rm, "p90": rp90, "sample": len(rs)}
    return rules


# ─── kunlik ma'lumotlar qurish ───
def build_dailydata(receipts, pnames, pskus, min_d, max_d, rules):
    days   = (max_d - min_d).days + 1
    labels = [(min_d.fromordinal(min_d.toordinal() + k)).isoformat() for k in range(days)]

    # ─── dinamik "oddiy mijoz" qutqarish parametrlari ───
    RESCUE_R       = 1.5   # qty < thr × R bo'lsa (mahsulotning o'z chegarasiga nisbatan biroz ko'p)
    RESCUE_BASKET  = 8     # savatda kamida shuncha xil tovar bo'lsa (keng savat = oddiy mijoz)

    item_data = {}
    for rc in receipts.values():
        di      = (rc["date"] - min_d).days
        explicit = is_wholesale(rc["customer"], rc["tin"])
        items    = rc["items"]
        # chek konteksti: savat kengligi va nechta tovar "ko'tarilgan" (qty>=thr)
        basket_n   = len(items)
        elevated_n = 0
        if not explicit:
            for _pk, _q in items.items():
                _r = rules.get(_pk)
                if _r and _q >= _r["thr"]:
                    elevated_n += 1
        for pk, qty in items.items():
            it = item_data.setdefault(pk, {
                "sku": pskus.get(pk, ""),
                "q": [0.0]*days, "r": [0]*days,
                "x": [0.0]*days, "i": [0.0]*days,
                "rr": [0]*days,  "wr": [0]*days,
                "rev_d": [0.0]*days,
                "revenue": 0.0,
                "exp_rec": 0, "inf_rec": 0,
                "ws_cust": defaultdict(float),
            })
            _rv = rc["item_rev"].get(pk, 0)
            it["q"][di]     += qty
            it["r"][di]     += 1
            it["rev_d"][di] += _rv
            it["revenue"]   += _rv
            if explicit:
                it["x"][di]  += qty
                it["wr"][di] += 1
                it["exp_rec"] += 1
                it["ws_cust"][rc["customer"] or rc["tin"] or "Korporativ"] += qty
                continue
            rule = rules.get(pk, {"cap": qty, "thr": float("inf"), "med": qty, "p90": qty, "sample": 0})
            if qty >= rule["thr"]:
                # DINAMIK qutqarish: mahsulotning o'z chegarasidan biroz oshgan (thr×R dan kam) +
                # chekда yolg'iz ko'tarilgan + keng savat → oddiy mijoz (bayram/uy xaridi) → RETAIL
                if qty < rule["thr"] * RESCUE_R and elevated_n == 1 and basket_n >= RESCUE_BASKET:
                    it["rr"][di] += 1
                else:
                    it["i"][di]  += max(0, qty - rule["cap"])
                    it["wr"][di] += 1
                    if rule["cap"] > 0: it["rr"][di] += 1
                    it["inf_rec"] += 1
            else:
                it["rr"][di] += 1

    output   = {}
    sku_al   = {}
    name_al  = {}

    for pk, it in item_data.items():
        rule   = rules.get(pk, {"cap": 0, "thr": 0, "med": 0, "p90": 0, "sample": 0})
        retail = [max(0, it["q"][d] - it["x"][d] - it["i"][d]) for d in range(days)]
        ws     = [it["x"][d] + it["i"][d] for d in range(days)]
        cname  = pnames[pk].most_common(1)[0][0]

        total  = sum(retail)
        active = sum(v > 0.001 for v in retail)
        fa     = total / days if days else 0
        a7     = sum(retail[-min(7,  days):]) / min(7,  days) if days >= 7  else fa
        a14    = sum(retail[-min(14, days):]) / min(14, days) if days >= 14 else fa
        if   days >= 21: daily = a7*0.5 + a14*0.3 + fa*0.2
        elif days >= 14: daily = a7*0.6 + fa*0.4
        else:            daily = fa

        f3  = sum(retail[:max(1, days//3)]) / max(1, days//3)
        l3  = sum(retail[-max(1, days//3):]) / max(1, days//3)
        if   f3 <= 0 and l3 > 0:        trend = "new"
        elif f3 > 0 and l3/f3 >= 1.25:  trend = "up"
        elif f3 > 0 and l3/f3 <= 0.75:  trend = "down"
        else:                            trend = "stable"

        obs    = sum(it["q"])
        ws_tot = sum(it["x"]) + sum(it["i"])
        conf   = round(100*(0.35*min(1.0,days/56) + 0.35*(active/days if days else 0) + 0.30*min(1.0,math.log10(total+1)/2)))

        m = {
            "daily":             round(daily, 2),
            "baselineDaily":     round(fa, 2),
            "week":              round(daily * 7, 2),
            "month":             round(daily * 30, 2),
            "calendarAvg":       round(fa, 2),
            "activeAvg":         round(total/active, 2) if active else 0,
            "activeDays":        active,
            "confidence":        max(0, min(100, conf)),
            "trend":             trend,
            "wholesalePct":      round(ws_tot/obs*100, 1) if obs else 0,
            "explicitWholesale": rq(sum(it["x"])),
            "inferredWholesale": rq(sum(it["i"])),
            "retailCap":         rq(rule["cap"]),
            "bulkThreshold":     rq(rule["thr"]),
            "receiptMedian":     rq(rule["med"]),
            "receiptP90":        rq(rule["p90"]),
            "receiptSample":     rule["sample"],
            "revenue":           round(it["revenue"]),
            "totalSold":         rq(sum(it["q"])),
            "totalReceipts":     sum(it["r"]),
            "explicitReceipts":  it["exp_rec"],
            "inferredReceipts":  it["inf_rec"],
            "wholesaleCustomers": [c for c,_ in sorted(it["ws_cust"].items(), key=lambda x:-x[1])[:3]],
        }

        output[pk] = {
            "name": cname, "sku": it["sku"],
            "q":  [rq(v) for v in it["q"]],
            "r":  it["r"], "rr": it["rr"], "wr": it["wr"],
            "w":  [rq(v) for v in ws],
            "x":  [rq(v) for v in it["x"]],
            "i":  [rq(v) for v in it["i"]],
            "rt": [rq(v) for v in retail],
            "rev": [round(v) for v in it["rev_d"]],
            "m":  m,
        }
        if it["sku"]:
            sku_al["sku:" + it["sku"]] = pk
        for n in pnames[pk]:
            name_al[n] = pk

    return {
        "__meta__": {
            "days": days, "start": min_d.isoformat(), "end": max_d.isoformat(),
            "title":  min_d.strftime("%B %Y"),
            "labels": labels,
            "method": "sku-customer-tin-dynamic-receipt-v3",
        },
        "items":       output,
        "skuAliases":  sku_al,
        "nameAliases": name_al,
    }


# ─── inventar ma'lumotlari ───
def build_invdata(products, arrivals=None):
    arrivals = arrivals or {}
    result = {}
    for sku, p in products.items():
        name = p["name"]
        if not name:
            continue
        entry = {
            "a":  rq(p["a"]),
            "sku": sku,
            "t":  p["tp"],
            "su": p["su"],
            "p":  rq(p["p"]),
            "sp": rq(p.get("sp", 0)),
            "sb": p["sub"],
            "cat": p.get("cat", ""),
            "catTop": p.get("catTop", "") or p.get("cat", ""),
        }
        arr = arrivals.get(str(sku))
        if arr and arr.get("date"):
            entry["la"] = arr["date"]
        result[name] = entry
    return result


# ─── savat tahlili ───
def build_basket(receipts, pk_set):
    prod_rc   = Counter()
    co_count  = defaultdict(Counter)

    for rc in receipts.values():
        pks = [pk for pk in rc["items"] if pk in pk_set]
        for pk in pks:
            prod_rc[pk] += 1
        if len(pks) < 2:
            continue
        for pk in pks:
            for other in pks:
                if other != pk:
                    co_count[pk][other] += 1

    basket = {}
    for pk, others in co_count.items():
        total = prod_rc[pk]
        if not total:
            continue
        basket[pk] = [
            {"pk": opk, "c": round(cnt / total * 100)}
            for opk, cnt in others.most_common(10)
            if round(cnt / total * 100) >= 3
        ]
    return basket


# ─── p2data ───
def build_p2data(receipts, pnames, pskus, daily_data, products, min_d, max_d):
    days    = (max_d - min_d).days + 1
    items_d = daily_data["items"]
    labels  = daily_data["__meta__"]["labels"]

    total_rev = sum(it["m"]["revenue"] for it in items_d.values())
    sorted_pks = sorted(items_d, key=lambda k: -items_d[k]["m"]["revenue"])

    # ABC klassifikatsiyasi
    cumrev  = 0
    abc_map = {}
    for pk in sorted_pks:
        cumrev += items_d[pk]["m"]["revenue"]
        pct     = cumrev / total_rev if total_rev else 0
        abc_map[pk] = "A" if pct <= 0.80 else ("B" if pct <= 0.95 else "C")

    basket    = build_basket(receipts, set(items_d.keys()))
    pk_to_name = {pk: pnames[pk].most_common(1)[0][0] for pk in items_d}

    # oxirgi sotuv sanasi + o'shandan o'tgan kunlar (di)
    last_sale = {}
    last_di   = {}
    for pk, it in items_d.items():
        for d in range(days - 1, -1, -1):
            if it["q"][d] > 0:
                last_sale[pk] = labels[d]
                last_di[pk]   = days - 1 - d
                break

    result = []
    for rank, pk in enumerate(sorted_pks, 1):
        it   = items_d[pk]
        sku  = pskus.get(pk, "")
        prod = products.get(sku, {})
        name = pnames[pk].most_common(1)[0][0]
        is_kg = any(kw in prod.get("tp", "").lower() for kw in ("кг", "kg", "кило"))

        b_out = [
            {"n": pk_to_name[e["pk"]], "c": e["c"]}
            for e in basket.get(pk, [])
            if e["pk"] in pk_to_name
        ]

        result.append({
            "r":    rank,
            "rev":  it["m"]["revenue"],
            "rp":   round(it["m"]["revenue"] / total_rev * 100, 2) if total_rev else 0,
            "qty":  it["m"]["totalSold"],
            "rec":  it["m"]["totalReceipts"],
            "p":    rq(prod.get("p", 0)),
            "kg":   is_kg,
            "ld":   last_sale.get(pk, ""),
            "di":   last_di.get(pk, days),
            "cat":  prod.get("cat", ""),
            "abc":  abc_map.get(pk, "C"),
            "b":    b_out,
            "d":    it["q"],
            "da":   it["m"]["daily"],   # aqlli kunlik velocity (retail + recency)
            "name": name,
            "sku":  sku,
        })

    return result


# ─── why/how matn generatsiyasi ───
def gen_why_how(abc, sub, di, active, days, trend, rev, qty, rec):
    day_avg = round(qty / days, 1) if days else 0

    if abc == "A":
        why = [
            "Daromadning 80% ini ta'minlovchi muhim mahsulot",
            f"Oyda {rec} ta chekda sotilgan — yuqori talab",
            f"Savdo {trend} tendensiyasida",
            "Zaxira tugashi butun savdoga zarar keltiradi",
        ]
        how = [
            "Zaxira hech qachon tugamasligini ta'minlash (safety stock oshirish)",
            "Yetkazib beruvchi bilan uzoq muddatli shartnoma tuzish",
            "Savdo hajmini haftalik monitoring qilish",
        ]
    elif abc == "B":
        why = [
            "Daromad ulushi 15% oralig'ida — o'rta muhimlikdagi mahsulot",
            f"Oyda {rec} ta chekda sotilgan",
            f"Savdo {trend} tendensiyasida",
            "A guruhiga o'tish imkoniyati mavjud",
        ]
        how = [
            "Savdo hajmini oshirish uchun A guruh mahsulotlari bilan birga taklif qilish",
            "Zaxira darajasini optimallashtirish — haddan oshiq buyurtma qilmaslik",
            "Aksiya vaqtida e'tibor berish — B dan A ga o'tkazish mumkin",
        ]
    else:
        if sub == "C1":
            why = [
                f"So'nggi {di} kun ichida savdo kuzatilmadi",
                f"Oylik daromad juda past ({round(rev/1000)}K so'm)",
                f"Jami {active} kun aktiv savdo bo'lgan ({days} kundan)",
                "Omborda qoldiq to'planib qolishi xavfi bor",
            ]
            how = [
                "Chegirma yoki aksiya bilan qolgan zaxirani sotish",
                "Yangi buyurtma to'xtatish",
                "30 kun ichida savdo bo'lmasa assortimentdan chiqarish",
            ]
        elif sub == "C2":
            why = [
                "Savdo hajmi pasayish tendensiyasida",
                f"Faqat {rec} ta chekda sotilgan (kam talab)",
                "Mijozlar boshqa alternativlarga o'tmoqda",
                "Daromad ulushi 5% dan past",
            ]
            how = [
                "Mahsulotni ko'p sotiluvchi mahsulotlar yonida joylash",
                "Narxni raqobatchilar bilan solishtirish",
                "Minimum zaxira darajasini kamaytirib, buyurtma hajmini qisqartirish",
            ]
        else:
            why = [
                "Mahsulot past chastotada, lekin barqaror sotiladi",
                f"O'rtacha {day_avg} dona/kun savdo (past hajm)",
                "Umumiy daromad ulushi 5% dan past",
                "Savdo barqaror lekin hajm kichik",
            ]
            how = [
                "Buyurtma hajmini minimal darajada ushlab turish",
                "Savat tahlili asosida ko'p sotiluvchi mahsulotlar yoniga joylashtirish",
                "Agar 2 oy ketma-ket C bo'lsa, assortiment qayta ko'rib chiqish",
            ]

    return why, how


# ─── p3data ───
def build_p3data(p2data, daily_data, max_d):
    days    = daily_data["__meta__"]["days"]
    items_d = daily_data["items"]

    # oxirgi sotuv kunidan bugungi kungacha o'tgan kunlar
    end_ord = max_d.toordinal()

    result = []
    for item in p2data:
        pk  = ("sku:" + item["sku"]) if item["sku"] else ("name:" + item["name"])
        it  = items_d.get(pk)

        # days since last sale
        di = days
        if it:
            for d in range(days - 1, -1, -1):
                if it["q"][d] > 0:
                    di = days - 1 - d
                    break

        active = it["m"]["activeDays"] if it else 0
        trend  = it["m"]["trend"]      if it else "stable"
        ws_pct = it["m"]["wholesalePct"] if it else 0

        abc = item["abc"]
        # C alt-klassifikatsiyasi
        if abc == "C":
            if di > 20:
                sub = "C1"
            elif trend == "down":
                sub = "C2"
            else:
                sub = "C3"
        else:
            sub = abc

        why, how = gen_why_how(
            abc, sub, di, active, days, trend,
            item["rev"], item["qty"], item["rec"]
        )

        result.append({
            "name": item["name"],
            "r":    item["r"],
            "rev":  item["rev"],
            "rp":   item["rp"],
            "qty":  item["qty"],
            "rec":  item["rec"],
            "p":    item["p"],
            "kg":   item["kg"],
            "ld":   item["ld"],
            "cat":  item["cat"],
            "abc":  abc,
            "sub":  sub,
            "di":   di,
            "tr":   trend,
            "why":  why,
            "how":  how,
        })

    return result


# ─── suppliers (P6) oylik ma'lumotlari ───
def build_supplier_months(month_rev, month_rec, month_name, products):
    """Har bir oy uchun ALOHIDA: o'sha oyning SKU bo'yicha daromadini (yengil,
    to'g'ridan-to'g'ri buyurtmalardan hisoblangan - og'ir kunlik/ulgurji
    pipeline'ga bog'liq emas) supplierga yig'ib, supplierlarni o'sha oyning
    daromadi bo'yicha qayta saralab ABC (80/15/5) belgilaydi. Bu orqali asosiy
    sahifalar (P1/P2/P3/kunlik/Zakas) qisqaroq oynada qolib, faqat Suppliers
    bo'limi butun tarixni (masalan 180 kun) ko'rib chiqishi mumkin.
    Natija: {supplier_nomi: {"YYYY-MM": {rev, rp, abc, cnt, rec, abc_cnt, top}}}."""
    result = defaultdict(dict)
    for month_key, sku_rev in month_rev.items():
        month_total = sum(sku_rev.values()) or 1
        sorted_skus = sorted(sku_rev, key=lambda k: -sku_rev[k])
        cum = 0.0
        sku_abc = {}
        for sku in sorted_skus:
            cum += sku_rev[sku]
            pct = cum / month_total
            sku_abc[sku] = "A" if pct <= 0.80 else ("B" if pct <= 0.95 else "C")

        rec_for_month = month_rec.get(month_key, {})
        sup_rev = defaultdict(float)
        sup_rec = defaultdict(int)
        sup_cnt = defaultdict(int)
        sup_abc_cnt = defaultdict(lambda: {"A": 0, "B": 0, "C": 0})
        sup_items = defaultdict(list)
        for sku, rev in sku_rev.items():
            supplier = products.get(sku, {}).get("su") or "Noma'lum"
            sup_rev[supplier] += rev
            sup_rec[supplier] += rec_for_month.get(sku, 0)
            sup_cnt[supplier] += 1
            sup_abc_cnt[supplier][sku_abc[sku]] += 1
            sup_items[supplier].append((sku, rev))

        total_rev = sum(sup_rev.values()) or 1
        sorted_sups = sorted(sup_rev, key=lambda n: -sup_rev[n])
        cum2 = 0.0
        for supplier in sorted_sups:
            cum2 += sup_rev[supplier]
            pct2 = cum2 / total_rev
            abc = "A" if pct2 <= 0.80 else ("B" if pct2 <= 0.95 else "C")
            top_items = sorted(sup_items[supplier], key=lambda e: -e[1])[:5]
            result[supplier][month_key] = {
                "rev": round(sup_rev[supplier]),
                "rp": round(sup_rev[supplier] / total_rev * 100, 2),
                "abc": abc,
                "cnt": sup_cnt[supplier],
                "rec": sup_rec[supplier],
                "abc_cnt": sup_abc_cnt[supplier],
                "top": [
                    {"name": month_name.get(sku, sku), "rev": round(rev),
                     "abc": sku_abc[sku], "sku": sku}
                    for sku, rev in top_items
                ],
            }
    return result


# ─── suppliers (P6) ma'lumotlari ───
def build_supplierdata(p2data, products, monthly=None, month_keys=None):
    """Har bir mahsulotni o'z ta'minotchisiga (supplier) bog'lab, supplier kesimida
    daromad/miqdor/chek/ABC taqsimotini hisoblaydi."""
    groups = defaultdict(lambda: {"rev": 0.0, "qty": 0.0, "rec": 0, "cnt": 0,
                                   "abc_cnt": {"A": 0, "B": 0, "C": 0}, "items": []})
    for item in p2data:
        supplier = products.get(item.get("sku", ""), {}).get("su") or "Noma'lum"
        group = groups[supplier]
        group["rev"] += item.get("rev", 0) or 0
        group["qty"] += item.get("qty", 0) or 0
        group["rec"] += item.get("rec", 0) or 0
        group["cnt"] += 1
        abc = item.get("abc", "C")
        if abc in group["abc_cnt"]:
            group["abc_cnt"][abc] += 1
        group["items"].append(item)

    total_rev = sum(g["rev"] for g in groups.values()) or 1
    sorted_names = sorted(groups, key=lambda n: -groups[n]["rev"])

    cumulative = 0.0
    suppliers = []
    abc_count = {"A": 0, "B": 0, "C": 0}
    for rank, name in enumerate(sorted_names, 1):
        group = groups[name]
        cumulative += group["rev"]
        pct = cumulative / total_rev
        abc = "A" if pct <= 0.80 else ("B" if pct <= 0.95 else "C")
        abc_count[abc] += 1
        top_items = sorted(group["items"], key=lambda it: -(it.get("rev", 0) or 0))[:5]
        entry = {
            "name": name,
            "rev": round(group["rev"]),
            "qty": rq(group["qty"]),
            "rec": group["rec"],
            "cnt": group["cnt"],
            "abc_cnt": group["abc_cnt"],
            "top": [
                {"name": it["name"], "rev": it.get("rev", 0), "abc": it.get("abc", "C"), "sku": it.get("sku", "")}
                for it in top_items
            ],
            "abc": abc,
            "r": rank,
            "rp": round(group["rev"] / total_rev * 100, 2),
        }
        if monthly is not None and month_keys is not None:
            sup_months = monthly.get(name, {})
            entry["months"] = [sup_months.get(mk) for mk in month_keys]
        suppliers.append(entry)

    return {
        "total_rev": round(total_rev),
        "sup_count": len(suppliers),
        "abc_cnt": abc_count,
        "suppliers": suppliers,
    }


# ─── sales.html ichiga joylashtirish ───
def embed_html(html_path, invdata, p2data, p3data, dailydata, p1data=None, supplierdata=None, template_path=None):
    # Shablon doim asosiy sales.html dan o'qiladi (formulalar o'zgarmaydi)
    src = template_path if template_path else html_path
    html = src.read_text(encoding="utf-8")

    def replace_block(sid, value):
        nonlocal html
        payload = json.dumps(value, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
        pat = re.compile(r'(<script[^>]+id="' + sid + r'"[^>]*>)[\s\S]*?(</script>)', re.I)
        html, n = pat.subn(lambda m: m.group(1) + payload + m.group(2), html, count=1)
        if n != 1:
            raise RuntimeError(f"Script blok topilmadi: {sid}")

    replace_block("invdata",   invdata)
    replace_block("p2data",    p2data)
    replace_block("p3data",    p3data)
    replace_block("dailydata", dailydata)
    if p1data is not None:
        replace_block("p1data", p1data)
    if supplierdata is not None:
        replace_block("supplierdata", supplierdata)

    html_path.write_text(html, encoding="utf-8")


# ─── asosiy funksiya ───
# ─── bosh sahifa (P1) ma'lumotlari ───
def build_p1data(receipts, pnames, pskus, pcats, refund_total, refund_by_day, p2data, products, min_d, max_d):
    days = (max_d - min_d).days + 1

    daily     = [0.0] * days
    daily_cost = [0.0] * days  # kunlik kelish narxi (tannarx)
    daily_rec = [0] * days
    weekly    = [0.0] * 7
    emp_rev   = defaultdict(float)
    emp_rec   = defaultdict(int)
    cat_rev   = defaultdict(float)
    emp_daily = defaultdict(lambda: [0.0] * days)   # xodim → kunlik tushum
    emp_rec_daily = defaultdict(lambda: [0] * days) # xodim → kunlik chek soni
    cat_daily = defaultdict(lambda: [0.0] * days)   # kategoriya → kunlik tushum
    gross     = 0.0
    gross_cost = 0.0

    sp_by_pk = {pk: products.get(sku, {}).get("sp", 0) or 0 for pk, sku in pskus.items()}

    for rc in receipts.values():
        di   = (rc["date"] - min_d).days
        rrev = sum(rc["item_rev"].values())
        rcost = sum(qty * sp_by_pk.get(pk, 0) for pk, qty in rc.get("items", {}).items())
        gross += rrev
        gross_cost += rcost
        in_range = 0 <= di < days
        if in_range:
            daily[di]     += rrev
            daily_cost[di] += rcost
            daily_rec[di] += 1
        weekly[rc["date"].weekday()] += rrev
        emp = rc["employee"] or "Noma'lum"
        emp_rev[emp] += rrev
        emp_rec[emp] += 1
        if in_range:
            emp_daily[emp][di] += rrev
            emp_rec_daily[emp][di] += 1
        for pk, rev in rc["item_rev"].items():
            if pcats.get(pk):
                cat = pcats[pk].most_common(1)[0][0]
            else:
                cat = products.get(pskus.get(pk, ""), {}).get("cat", "")
            if cat:
                cat_rev[cat] += rev
                if in_range:
                    cat_daily[cat][di] += rev

    gross_profit = gross - gross_cost

    nrec      = len(receipts)
    avg_check = gross / nrec if nrec else 0
    refund_pct = refund_total / gross * 100 if gross else 0
    sku_count = len(p2data)
    staff     = len([e for e in emp_rev if e and e != "Noma'lum"]) or len(emp_rev)

    # kun raqamlari (oy ichidagi sana)
    day_labels = [str((min_d.fromordinal(min_d.toordinal() + k)).day) for k in range(days)]

    weekly_out = [{"day": WEEKDAYS_UZ[i], "val": round(weekly[i])} for i in range(7)]
    top_cats   = [{"name": n, "val": round(v)} for n, v in sorted(cat_rev.items(), key=lambda x: -x[1])[:8]]

    def _item_cost_profit(it):
        sp = products.get(it.get("sku", ""), {}).get("sp", 0) or 0
        cost = sp * (it.get("qty", 0) or 0)
        rev = it.get("rev", 0) or 0
        return round(cost), round(rev - cost)

    top_items = []
    for it in p2data[:8]:
        cost, profit = _item_cost_profit(it)
        top_items.append({"name": it["name"], "val": it["rev"], "cost": cost, "profit": profit})

    # faqat kelish narxi Invan'da kiritilgan mahsulotlar - aks holda noma'lum
    # tannarx 0 deb olinib, sun'iy ravishda "100% foyda" bo'lib chiqib qoladi
    has_cost = [it for it in p2data if (products.get(it.get("sku", ""), {}).get("sp", 0) or 0) > 0]
    top_items_profit = []
    for it in sorted(has_cost, key=lambda it: _item_cost_profit(it)[1], reverse=True)[:8]:
        cost, profit = _item_cost_profit(it)
        top_items_profit.append({"name": it["name"], "val": profit, "rev": it.get("rev", 0), "cost": cost})

    top_emp    = [{"name": n, "val": round(emp_rev[n]), "rec": emp_rec[n]}
                  for n in sorted(emp_rev, key=lambda x: -emp_rev[x])[:8]]

    a_count = sum(1 for it in p2data if it["abc"] == "A")
    b_count = sum(1 for it in p2data if it["abc"] == "B")
    c_count = sum(1 for it in p2data if it["abc"] == "C")
    a_rev   = sum(it["rev"] for it in p2data if it["abc"] == "A")
    b_rev   = sum(it["rev"] for it in p2data if it["abc"] == "B")
    c_rev   = sum(it["rev"] for it in p2data if it["abc"] == "C")
    total_count = len(p2data) or 1

    best_i  = max(range(days), key=lambda i: daily[i]) if days else 0
    worst_i = min(range(days), key=lambda i: daily[i]) if days else 0

    # ── oraliq (date-range) filtri uchun kunlik massivlar ──
    iso_dates    = [(min_d.fromordinal(min_d.toordinal() + k)).isoformat() for k in range(days)]
    daily_refund = [round(refund_by_day.get(min_d.fromordinal(min_d.toordinal() + k), 0)) for k in range(days)]
    # barcha xodimlar (kichik ro'yxat) — kunlik tushum va chek soni
    emp_daily_out = {e: [round(x) for x in arr] for e, arr in emp_daily.items()}
    emp_rec_daily_out = {e: list(arr) for e, arr in emp_rec_daily.items()}
    # barcha kategoriyalar — kunlik tushum
    cat_daily_out = {c: [round(x) for x in arr] for c, arr in cat_daily.items()}
    # top 120 mahsulot — kunlik tushum va kelish narxi (rev/cost * kunlik_miqdor/jami_miqdor)
    items_daily = []
    for it in p2data[:120]:
        q = it.get("qty", 0) or 0
        rev = it.get("rev", 0) or 0
        sp = products.get(it.get("sku", ""), {}).get("sp", 0) or 0
        dq = it.get("d", [])
        if q > 0 and dq:
            items_daily.append({
                "name": it["name"],
                "d": [round(rev * (x / q)) for x in dq],
                "c": [round(sp * x) for x in dq],
                "hc": sp > 0,  # kelish narxi Invan'da kiritilganmi (foyda reytingi uchun)
            })

    mname = MONTHS_UZ.get(min_d.month, str(min_d.month))
    title = f"{mname} {min_d.year}"
    if min_d.month == max_d.month and min_d.year == max_d.year:
        period_text = (f"{min_d.year}-yil {mname.lower()} oyi · {min_d.day}-{max_d.day} {mname.lower()} · "
                       f"{days} kunlik ma'lumot")
    else:
        period_text = f"{min_d.isoformat()} — {max_d.isoformat()} · {days} kunlik ma'lumot"

    return {
        "title":       title,
        "periodText":  period_text,
        "days":        days,
        "start":       min_d.isoformat(),
        "end":         max_d.isoformat(),
        "gross":       round(gross),
        "cost":        round(gross_cost),
        "profit":      round(gross_profit),
        "refund":      round(refund_total),
        "refund_pct":  round(refund_pct, 2),
        "receipts":    nrec,
        "avg_check":   round(avg_check),
        "sku":         sku_count,
        "staff":       staff,
        "daily":       [round(v) for v in daily],
        "dailyCost":   [round(v) for v in daily_cost],
        "dayLabels":   day_labels,
        "weekly":      weekly_out,
        "top_cats":    top_cats,
        "top_items":   top_items,
        "top_items_profit": top_items_profit,
        "top_emp":     top_emp,
        "dates":       iso_dates,
        "dailyRec":    daily_rec,
        "dailyRefund": daily_refund,
        "empDaily":    emp_daily_out,
        "empRecDaily": emp_rec_daily_out,
        "catDaily":    cat_daily_out,
        "itemsDaily":  items_daily,
        "abc": {
            "a_count": a_count, "b_count": b_count, "c_count": c_count,
            "a_rev": round(a_rev), "b_rev": round(b_rev), "c_rev": round(c_rev),
        },
        "c_assort_pct": round(c_count / total_count * 100),
        "best_day":  {"idx": best_i,  "label": day_labels[best_i]  if days else "",
                      "val": round(daily[best_i])  if days else 0},
        "worst_day": {"idx": worst_i, "label": day_labels[worst_i] if days else "",
                      "val": round(daily[worst_i]) if days else 0},
    }


def build(sales_path, products_path, html_path=None):
    if html_path is None:
        html_path = ROOT / "sales.html"

    print(f"[1/6] Tovarlar o'qilmoqda: {Path(products_path).name}")
    products = read_products(products_path)
    print(f"      {len(products):,} mahsulot")

    print(f"[2/6] Sotuv o'qilmoqda: {Path(sales_path).name}")
    receipts, pnames, pskus, pcats, refund_total, refund_by_day, min_d, max_d = read_sales(sales_path)
    print(f"      {len(receipts):,} chek  |  {min_d} — {max_d}")

    print(f"[3/6] Ulgurji qoidalar hisoblanmoqda...")
    rules = compute_rules(receipts)
    print(f"      {len(rules):,} mahsulot uchun chegara aniqlandi")

    print(f"[4/6] Kunlik tahlil qurilmoqda...")
    dailydata = build_dailydata(receipts, pnames, pskus, min_d, max_d, rules)
    print(f"      {len(dailydata['items']):,} mahsulot")

    print(f"[5/6] Mahsulot, inventar va ABC ma'lumotlari qurilmoqda...")
    invdata = build_invdata(products)
    p2data  = build_p2data(receipts, pnames, pskus, dailydata, products, min_d, max_d)
    p3data  = build_p3data(p2data, dailydata, max_d)
    p1data  = build_p1data(receipts, pnames, pskus, pcats, refund_total, refund_by_day, p2data, products, min_d, max_d)
    supplierdata = build_supplierdata(p2data, products)
    A = sum(1 for i in p2data if i["abc"] == "A")
    B = sum(1 for i in p2data if i["abc"] == "B")
    C = sum(1 for i in p2data if i["abc"] == "C")
    print(f"      ABC: A={A}, B={B}, C={C}")

    # data_daily.json ni ham yangilash — API server shu fayldan o'qiydi
    # API server shu uchta fayldan o'qiydi — ularni ham yangilash shart
    def _write_json(name, data):
        (ROOT / name).write_text(
            json.dumps(data, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8"
        )
    _write_json("data_daily.json",       dailydata)   # talab/kunlik
    _write_json("data_mahsulotlar.json", p2data)      # mahsulot kartochkalari (P2)
    _write_json("data_inv_new.json",     invdata)     # inventar/qoldiq
    _write_json("data_supplier.json",    supplierdata)  # ta'minotchilar (P6)

    print(f"[6/6] sales.html yangilanmoqda: {html_path.name}")
    template = ROOT / "sales.html"
    embed_html(html_path, invdata, p2data, p3data, dailydata, p1data, supplierdata,
               template_path=template if template != html_path else None)

    # Vercel uchun: index.html ni sales.html dan nusxalash (root sahifa)
    index_path = ROOT / "index.html"
    try:
        shutil.copyfile(html_path, index_path)
        print(f"      index.html yangilandi (Vercel uchun)")
    except Exception as e:
        print(f"      ! index.html nusxalashda xato: {e}")

    title = min_d.strftime("%B %Y")
    return {
        "status":   "ok",
        "period":   title,
        "products": len(p2data),
        "receipts": len(receipts),
        "abc":      {"A": A, "B": B, "C": C},
    }


def main():
    ap = argparse.ArgumentParser(description="Yangi oy ma'lumotlari bilan sales.html ni yangilash")
    ap.add_argument("--sales",    default="sotuv_excel.xlsx",  help="Sotuv Excel fayli")
    ap.add_argument("--products", default="Товары (6).xlsx",   help="Tovarlar Excel fayli")
    ap.add_argument("--output",   default="sales.html",         help="HTML chiqish fayli")
    args = ap.parse_args()

    result = build(ROOT / args.sales, ROOT / args.products, ROOT / args.output)

    print(f"\n{'='*40}")
    print(f"  TAYYOR! {result['period']}")
    print(f"  Mahsulotlar:  {result['products']:,}")
    print(f"  Cheklar:      {result['receipts']:,}")
    print(f"  ABC: A={result['abc']['A']}, B={result['abc']['B']}, C={result['abc']['C']}")
    print(f"{'='*40}")


if __name__ == "__main__":
    main()
